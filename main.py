import openpyxl as xl
import networkx as nx
from get_in_out_degree import getInDegree
from get_in_out_degree import getOutDegree
from get_in_out_degree import getPerformance
from matplotlib import pyplot as plt

match_1_sheet = xl.load_workbook('data MATH 225.xlsx')
sheet1 = match_1_sheet['Sheet1']


match_2_sheet = xl.load_workbook('data match2.xlsx')
sheet2 = match_2_sheet['Sheet1']

def loadGraph(position_x, position_y,number_of_nodes, sheet):
    new_graph = nx.DiGraph()
    for i in range(1, number_of_nodes+1):
        new_graph.add_node(i)
    for i in range(0,len(new_graph.nodes(data=True))):
        data = list(new_graph.nodes(data=True))[i][1]
        data['name'] = sheet.cell( position_x + i, position_y-1).value
        data['position'] = sheet.cell( position_x + i, position_y-2).value
        data['number'] = sheet.cell( position_x + i, position_y-3).value

    for j in range(position_y, position_y+number_of_nodes): #to go through columns
        for i in range(position_x, position_x+number_of_nodes): #to through rows
            if(sheet.cell(i,j).value != 0):
                new_graph.add_edge(i - position_x + 1, j - position_y + 1, weight=sheet.cell(i,j).value)


    return new_graph


#match1:

match1_0_10 = loadGraph(9, 23, 11, sheet1)
match1_10_20 = loadGraph(25, 23, 11, sheet1)
match1_20_30 = loadGraph(41, 23, 11, sheet1)
match1_30_40 = loadGraph(57, 23, 11, sheet1)
match1_40_50 = loadGraph(73, 23, 11, sheet1)
match1_50_60 = loadGraph(89, 23, 11, sheet1)
match1_60_70 = loadGraph(105, 23, 11, sheet1)
match1_70_80 = loadGraph(121, 23, 13, sheet1)
match1_80_90 = loadGraph(139, 23, 14, sheet1)
match1_overtime = loadGraph(158, 23, 14, sheet1)

#match2:

match2_0_10 = loadGraph(8, 6, 11, sheet2)
match2_10_20 = loadGraph(25, 6, 11, sheet2)
match2_20_30 = loadGraph(39, 6, 11, sheet2)
match2_30_40 = loadGraph(53, 6, 11, sheet2)
match2_40_50 = loadGraph(66, 6, 12, sheet2)
match2_50_60 = loadGraph(82, 6, 12, sheet2)
match2_60_70 = loadGraph(98, 6, 14, sheet2)
match2_70_80 = loadGraph(116, 6, 14, sheet2)
match2_80_90 = loadGraph(134, 6, 14, sheet2)
match2_overtime = loadGraph(151, 6, 14, sheet2)



performances_match_1 = []
performances_match_2 = []
#match 1 performance
print('match1 performance:\n')
print ('from 0 to 10 ')
performances_match_1.append(getPerformance(match1_0_10))
print ('\n from 10 to 20 ')
performances_match_1.append(getPerformance(match1_10_20))
print ('\n from 20 to 30:')
performances_match_1.append(getPerformance(match1_20_30))
print ('\n from 30 to 40:')
performances_match_1.append(getPerformance(match1_30_40))
print ('\n from 40 to 50:')
performances_match_1.append(getPerformance(match1_40_50))
print ('\n from 50 to 60:')
performances_match_1.append(getPerformance(match1_50_60))
print ('\n from 60 to 70:')
performances_match_1.append(getPerformance(match1_60_70))
print ('\n from 70 to 80:')
performances_match_1.append(getPerformance(match1_70_80))
print ('\n from 80 to 90:')
performances_match_1.append(getPerformance(match1_80_90))
print ('\n overtime:')
performances_match_1.append(getPerformance(match1_overtime))
print ('\n\n')

#match 2 performance
print('match2 performance:\n')
print ('from 0 to 10 ')
performances_match_2.append(getPerformance(match2_0_10))
print ('\n from 10 to 20 ')
performances_match_2.append(getPerformance(match2_10_20))
print ('\n from 20 to 30:')
performances_match_2.append(getPerformance(match2_20_30))
print ('\n from 30 to 40:')
performances_match_2.append(getPerformance(match2_30_40))
print ('\n from 40 to 50:')
performances_match_2.append(getPerformance(match2_40_50))
print ('\n from 50 to 60:')
performances_match_2.append(getPerformance(match2_50_60))
print ('\n from 60 to 70:')
performances_match_2.append(getPerformance(match2_60_70))
print ('\n from 70 to 80:')
performances_match_2.append(getPerformance(match2_70_80))
print ('\n from 80 to 90:')
performances_match_2.append(getPerformance(match2_80_90))
print ('\n overtime:')
performances_match_2.append(getPerformance(match2_overtime))


def findTotalPerformance(overall_performances, property = 'name'):
    total_performance = {}
    for performances in overall_performances:
        for performance in performances:
            name = list(performance.items())[0][0]
            perform = list(performance.items())[0][1]
            if name in total_performance:
                total_performance[name] += perform
            else:
                total_performance[name] = perform
    return total_performance

def printPerformances(performances):
    for k, v in performances.items():
        print(f'{k}: {v}')



print('\n\n')
print ('total performance for match 1 :\n')
printPerformances(findTotalPerformance(performances_match_1))
print('\n\ntotal performance for match 2 :\n')
printPerformances(findTotalPerformance(performances_match_2))
print('\n')

print ('game 1 clustering:\n')
print(nx.cluster.clustering(match1_0_10))
print(nx.cluster.clustering(match1_10_20))
print(nx.cluster.clustering(match1_20_30))
print(nx.cluster.clustering(match1_30_40))
print(nx.cluster.clustering(match1_40_50))
print(nx.cluster.clustering(match1_50_60))
print(nx.cluster.clustering(match1_60_70))
print(nx.cluster.clustering(match1_70_80))
print(nx.cluster.clustering(match1_80_90))
print(nx.cluster.clustering(match1_overtime))

print('\n\n match 2 clustering:\n')
print(nx.cluster.clustering(match2_0_10))
print(nx.cluster.clustering(match2_10_20))
print(nx.cluster.clustering(match2_20_30))
print(nx.cluster.clustering(match2_30_40))
print(nx.cluster.clustering(match2_40_50))
print(nx.cluster.clustering(match2_50_60))
print(nx.cluster.clustering(match2_60_70))
print(nx.cluster.clustering(match2_70_80))
print(nx.cluster.clustering(match2_80_90))
print(nx.cluster.clustering(match2_overtime))

print('\n match 1 pagerank:\n')
print(nx.pagerank(match1_0_10))
print(nx.pagerank(match1_10_20))
print(nx.pagerank(match1_20_30))
print(nx.pagerank(match1_30_40))
print(nx.pagerank(match1_40_50))
print(nx.pagerank(match1_50_60))
print(nx.pagerank(match1_60_70))
print(nx.pagerank(match1_70_80))
print(nx.pagerank(match1_80_90))
print(nx.pagerank(match1_overtime))

print('\n\n match 2 pagerank:\n')
print(nx.pagerank(match2_0_10))
print(nx.pagerank(match2_10_20))
print(nx.pagerank(match2_20_30))
print(nx.pagerank(match2_30_40))
print(nx.pagerank(match2_40_50))
print(nx.pagerank(match2_50_60))
print(nx.pagerank(match2_60_70))
print(nx.pagerank(match2_70_80))
print(nx.pagerank(match2_80_90))
print(nx.pagerank(match2_overtime))

print('\n\n match 1 betweenness_centrality:\n')
print(nx.betweenness_centrality(match1_0_10))
print(nx.betweenness_centrality(match1_10_20))
print(nx.betweenness_centrality(match1_20_30))
print(nx.betweenness_centrality(match1_30_40))
print(nx.betweenness_centrality(match1_40_50))
print(nx.betweenness_centrality(match1_50_60))
print(nx.betweenness_centrality(match1_60_70))
print(nx.betweenness_centrality(match1_70_80))
print(nx.betweenness_centrality(match1_70_80))
print(nx.betweenness_centrality(match1_70_80))
print(nx.betweenness_centrality(match1_80_90))
print(nx.betweenness_centrality(match1_overtime))

print('\n\n match 2 betweenness_centrality:\n')
print(nx.betweenness_centrality(match2_0_10))
print(nx.betweenness_centrality(match2_10_20))
print(nx.betweenness_centrality(match2_20_30))
print(nx.betweenness_centrality(match2_30_40))
print(nx.betweenness_centrality(match2_40_50))
print(nx.betweenness_centrality(match2_50_60))
print(nx.betweenness_centrality(match2_60_70))
print(nx.betweenness_centrality(match2_70_80))
print(nx.betweenness_centrality(match2_80_90))
print(nx.betweenness_centrality(match2_overtime))

def getLabels(graph, property = 'name'):
    labels = {}
    for node in graph.nodes(data=True):
        labels[node[0]] = node[1][property]
    return labels

#match 1 garphs:

nx.draw_circular(match1_0_10, labels = getLabels(match1_0_10, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match1_10_20, labels = getLabels(match1_10_20, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match1_20_30, labels = getLabels(match1_20_30, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match1_30_40, labels = getLabels(match1_30_40, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match1_40_50, labels = getLabels(match1_40_50, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match1_50_60, labels = getLabels(match1_50_60, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match1_60_70, labels = getLabels(match1_60_70, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match1_70_80, labels = getLabels(match1_70_80, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match1_80_90, labels = getLabels(match1_80_90, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match1_overtime, labels = getLabels(match1_overtime, 'number'), with_labels=True)
plt.show()

#match 2 graphs:

nx.draw_circular(match2_0_10, labels = getLabels(match2_0_10, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match2_10_20, labels = getLabels(match2_10_20, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match2_20_30, labels = getLabels(match2_20_30, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match2_30_40, labels = getLabels(match2_30_40, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match2_40_50, labels = getLabels(match2_40_50, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match2_50_60, labels = getLabels(match2_50_60, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match2_60_70, labels = getLabels(match2_60_70, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match2_70_80, labels = getLabels(match2_70_80, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match2_80_90, labels = getLabels(match2_80_90, 'number'), with_labels=True)
plt.show()
nx.draw_circular(match2_overtime, labels = getLabels(match2_overtime, 'number'), with_labels=True)
plt.show()